import os

from openpyxl import load_workbook

class Excel:
    def __init__(self, filename):
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        self.__name = os.path.join("data", filename)


    def get_name(self) -> str:
        return self.__name

    def get_data(self):
        if not self.check_excel_exists():
            return None

        lines = []
        try:
            wb = load_workbook(self.__name)
            ws = wb.active

            for i in range(2, ws.max_row + 1):
                position_text = f"A{i}"
                position_emotion = f"B{i}"

                text = ws[position_text].internal_value
                emotion = ws[position_emotion].value

                lines.append((text, emotion))

            wb.close()
        except Exception as e:
            print(e)

        return lines

    def check_excel_exists(self):
        if os.path.exists(self.get_name()):
            return True
        else:
            print(f"File: {self.get_name()} does not exist")
            return False